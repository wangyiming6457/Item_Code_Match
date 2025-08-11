import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from rapidfuzz import fuzz
from io import BytesIO
from copy import copy

# --- AUTH ---
def login():
    st.title("🔒 Cost Sheet Login")
    username = st.text_input("Username")
    password = st.text_input("Password", type="password")
    if st.button("Login"):
        if username == "ademco" and password == "yimingiscool":
            st.session_state.logged_in = True
        else:
            st.error("Invalid username or password.")

if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

if not st.session_state.logged_in:
    login()
    st.stop()

# --- Utilities ---
def clean_model(text):
    if not isinstance(text, str):
        return ""
    return (
        text.strip()
        .lower()
        .replace(" ", "")
        .replace("-", "")
        .replace("\n", "")
        .split("(")[0]
    )

def find_header_row_and_model_col(sheet):
    for row in sheet.iter_rows(min_row=1, max_row=40):
        for cell in row:
            if isinstance(cell.value, str) and "model no" in cell.value.lower():
                return cell.row, cell.column
    return None, None

# --- Streamlit UI ---
st.title("⚡ Cost Sheet Item Code Matcher (First Sheet Only)")
st.write("Upload your Item Listing and Cost Sheet Excel files.")

item_file = st.file_uploader("Upload Item Listing Excel", type="xlsx")
cost_file = st.file_uploader("Upload Cost Sheet Excel (first sheet will be used)", type="xlsx")

if item_file and cost_file:
    with st.spinner("Processing first sheet... Please wait..."):
        # Load item listing
        item_listing_df = pd.read_excel(item_file)
        item_lookup = {
            clean_model(name): (name, code)
            for name, code in zip(item_listing_df['Display Name'], item_listing_df['Name'])
            if isinstance(name, str) and clean_model(name) != ""
        }

        # Load cost workbook (first sheet only)
        original_wb = load_workbook(cost_file)
        first_sheet_name = original_wb.sheetnames[0]
        sheet = original_wb[first_sheet_name]

        # Prepare output workbook
        output_wb = Workbook()
        output_wb.remove(output_wb.active)
        new_sheet = output_wb.create_sheet(title=first_sheet_name)

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        header_row, model_col = find_header_row_and_model_col(sheet)
        if not header_row or not model_col:
            st.error("Model No column not found in the first sheet.")
            st.stop()

        # Copy data without heavy style copying
        for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row, max_row=sheet.max_row), start=1):
            for col_idx, cell in enumerate(row, start=1):
                new_sheet.cell(row=row_idx, column=col_idx, value=cell.value)

        # Insert Item Code column
        insert_at = model_col
        new_sheet.insert_cols(insert_at)
        new_sheet.cell(row=1, column=insert_at).value = "Item Code"

        # Matching logic
        for row in range(2, new_sheet.max_row + 1):
            raw_model = new_sheet.cell(row=row, column=insert_at + 1).value
            if not isinstance(raw_model, str) or not raw_model.strip():
                continue

            cleaned_model = clean_model(raw_model)
            if cleaned_model == "" or cleaned_model.isdigit():
                continue

            matched_code = None

            # Exact match
            if cleaned_model in item_lookup:
                matched_code = item_lookup[cleaned_model][1]
            else:
                # Fuzzy match
                best_score = 0
                best_code = None
                for key, (disp, code) in item_lookup.items():
                    score = fuzz.token_set_ratio(cleaned_model, key)
                    if score > best_score:
                        best_score = score
                        best_code = code
                if best_score >= 90:
                    matched_code = best_code
                else:
                    # Prefix match
                    for key, (disp, code) in item_lookup.items():
                        if cleaned_model.startswith(key):
                            matched_code = code
                            break

            item_cell = new_sheet.cell(row=row, column=insert_at)
            if matched_code:
                item_cell.value = matched_code
            else:
                item_cell.fill = yellow_fill

        # Prepare download
        output = BytesIO()
        output_wb.save(output)
        st.success("✅ Done! Download your processed file below.")
        st.download_button(
            label="📥 Download Updated Cost Sheet",
            data=output.getvalue(),
            file_name="CostSheet_ItemCode_Matched.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


